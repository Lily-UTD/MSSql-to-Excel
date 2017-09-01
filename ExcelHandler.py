# pip install openpyxl

import openpyxl
import os.path

class excelhandler(object):
    def __init__(self, filename,sheetname):
        self.filename = filename
        if os.path.exists(filename):
            self.wb = openpyxl.load_workbook(self.filename)
            self.asheet = self.wb.create_sheet(sheetname)
        else:
            self.wb = openpyxl.Workbook()
            self.asheet = self.wb.active
            self.asheet.title = sheetname
        self.thin_border = openpyxl.styles.borders.Border(left = openpyxl.styles.borders.Side(style='thin'), 
                     right = openpyxl.styles.borders.Side(style='thin'), 
                     top = openpyxl.styles.borders.Side(style='thin'), 
                     bottom = openpyxl.styles.borders.Side(style='thin'))
    
    def setHeader(self, header):
        col=1
        if len(header) > 0:
            for head in header:
                self.asheet.cell(row=1, column=col).value = head
                self.asheet.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
                self.asheet.cell(row=1, column=col).border = self.thin_border
                col = col+1
    
    def setData(self,data):
        numrow = 2
        if len(data) > 0:
            for row in data:
                if len(row) > 0:
                    numcol = 1
                    for cell in row:
                        self.asheet.cell(row=numrow, column=numcol).value = cell
                        self.asheet.cell(row=numrow, column=numcol).border = self.thin_border
                        numcol = numcol + 1
                    numrow = numrow + 1
    
    def saveExcl(self):
        self.wb.save(self.filename)